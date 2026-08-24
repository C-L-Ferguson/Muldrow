"""
Muldrow v. City of St. Louis — Empirical Analysis
Steps 1–10: Descriptive stats, pre/post comparisons, regressions, visualizations
"""

import warnings
warnings.filterwarnings("ignore")

import re
import numpy as np
import pandas as pd
from scipy import stats
from scipy.stats import chi2_contingency
import statsmodels.formula.api as smf
from statsmodels.miscmodels.ordinal_model import OrderedModel
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ── Paths ─────────────────────────────────────────────────────────────────────
INFILE  = "/root/.claude/uploads/39baeedb-fb1d-5fba-9c0a-9a900b672d1e/cf0bb9b3-muldrow_coding_sheet_cleaned.xlsx"
OUTDIR  = "/home/user/Muldrow"
XLOUT   = os.path.join(OUTDIR, "muldrow_tables.xlsx")
MULDROW_DATE = pd.Timestamp("2024-04-17")

# ── Colour palette (publication-quality, colourblind-friendly) ────────────────
BLUE    = "#2E6DA4"
ORANGE  = "#E07B39"
GREEN   = "#3A8C5C"
RED     = "#C0392B"
PURPLE  = "#7B5EA7"
GREY    = "#7F8C8D"

PAL_2   = [BLUE, ORANGE]
PAL_3   = [BLUE, ORANGE, GREEN]
PAL_4   = [BLUE, ORANGE, GREEN, RED]

plt.rcParams.update({
    "font.family":      "DejaVu Sans",
    "font.size":        10,
    "axes.titlesize":   11,
    "axes.titleweight": "bold",
    "axes.spines.top":  False,
    "axes.spines.right":False,
    "figure.dpi":       300,
    "savefig.dpi":      300,
    "savefig.bbox":     "tight",
})

# ═════════════════════════════════════════════════════════════════════════════
# 0. LOAD & CLEAN
# ═════════════════════════════════════════════════════════════════════════════
df_raw = pd.read_excel(INFILE, sheet_name="Coding Sheet", header=0)
print(f"Loaded {len(df_raw)} rows")

# Standardise key column names
df = df_raw.copy()

# Parse dates
df["Filing Date"]   = pd.to_datetime(df["Filing Date"],   errors="coerce")
df["Decision Date"] = pd.to_datetime(df["Decision Date"], errors="coerce")

# Drop rows with missing Period or Court Finding on Harm
df = df.dropna(subset=["Period (Pre/Post)"])
print(f"After dropping blank Period: {len(df)} rows")

# Normalise "Causation / Discriminatory Intent Insufficient" → short label
DISMISSAL_MAP = {
    "Harm Insufficient":                              "Harm Insufficient",
    "Causation / Discriminatory Intent Insufficient": "Causation Insufficient",
    "N/A — Plaintiff Survived":                       "N/A — Plaintiff Survived",
    "Pleading Deficiency":                            "Pleading Deficiency",
    "Procedural / Forfeiture":                        "Other",
    "Multiple Bases":                                 "Other",
    "Other":                                          "Other",
}
df["Dismissal_Clean"] = df["Primary Basis for Dismissal"].map(DISMISSAL_MAP).fillna("Other")

HARM_ORDER      = ["Insufficient", "Sufficient", "Mixed", "Not Reached"]
DISMISSAL_ORDER = ["Harm Insufficient", "Causation Insufficient",
                   "N/A — Plaintiff Survived", "Pleading Deficiency", "Other"]

# Filing-date subgroup for post-period
def filing_subgroup(row):
    if row["Period (Pre/Post)"] == "Pre":
        return "Pre-Period"
    fd = row["Filing Date"]
    if pd.isna(fd):
        return "Post — Pre-Muldrow Filing"   # assume pre if unknown
    return ("Post — Post-Muldrow Filing" if fd >= MULDROW_DATE
            else "Post — Pre-Muldrow Filing")

df["Filing_Subgroup"] = df.apply(filing_subgroup, axis=1)

pre  = df[df["Period (Pre/Post)"] == "Pre"].copy()
post = df[df["Period (Pre/Post)"] == "Post"].copy()
print(f"Pre: {len(pre)}  Post: {len(post)}")

# ═════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def pct(n, total):
    return f"{n} ({100*n/total:.1f}%)" if total > 0 else "—"

def chi2_result(ct):
    """Return chi2, p-value string from a contingency table (DataFrame or array)."""
    arr = ct.values if isinstance(ct, pd.DataFrame) else np.array(ct)
    # Drop all-zero rows/cols
    arr = arr[arr.sum(axis=1) > 0][:, arr.sum(axis=0) > 0]
    if arr.shape[0] < 2 or arr.shape[1] < 2:
        return np.nan, "—"
    chi2, p, dof, _ = chi2_contingency(arr, correction=False)
    stars = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""
    return chi2, f"{p:.3f}{stars}"

def crosstab_pct(data, row_col, col_col, row_order=None, col_order=None):
    ct  = pd.crosstab(data[row_col], data[col_col])
    if row_order:
        ct = ct.reindex([r for r in row_order if r in ct.index])
    if col_order:
        ct = ct.reindex(columns=[c for c in col_order if c in ct.columns], fill_value=0)
    pct_ct = ct.div(ct.sum(axis=0), axis=1) * 100
    return ct, pct_ct

def sig_label(p):
    if np.isnan(p): return ""
    return "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""

# ═════════════════════════════════════════════════════════════════════════════
# EXCEL WORKBOOK SETUP
# ═════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

HDR_FILL   = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL= PatternFill("solid", fgColor="2E75B6")
ALT_FILL   = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
SUBHDR_FONT= Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BODY_FONT  = Font(name="Calibri", size=10)
BOLD_FONT  = Font(bold=True, name="Calibri", size=10)

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="1F3864")

def thin_border(top=False, bottom=False, left=False, right=False):
    return Border(
        top=THIN if top else Side(style=None),
        bottom=THIN if bottom else Side(style=None),
        left=THIN if left else Side(style=None),
        right=THIN if right else Side(style=None),
    )

def add_sheet(title):
    ws = wb.create_sheet(title=title[:31])
    return ws

def write_header(ws, row, cols, merge_end=None, fill=HDR_FILL, font=HDR_FONT):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    if merge_end:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=merge_end)

def write_row(ws, row, vals, bold=False, fill=None, indent=0):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BOLD_FONT if bold else BODY_FONT
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(
            horizontal="left" if c == 1 else "center",
            indent=indent if c == 1 else 0
        )
        cell.border = thin_border(bottom=True)

def autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max_w, max(min_w, max_len + 2))

def add_note(ws, row, note):
    cell = ws.cell(row=row, column=1, value=note)
    cell.font = Font(italic=True, color="595959", size=9)

# ═════════════════════════════════════════════════════════════════════════════
# STEP 1 — DESCRIPTIVE STATISTICS TABLE
# ═════════════════════════════════════════════════════════════════════════════
print("\n── Step 1: Descriptive Statistics ──")

ws1 = add_sheet("Table 1 — Descriptive Stats")
ws1.freeze_panes = "B3"

N_total = len(df)
N_pre   = len(pre)
N_post  = len(post)

# Header
ws1.merge_cells("A1:D1")
ws1["A1"] = "Table 1. Sample Characteristics by Period"
ws1["A1"].font = Font(bold=True, size=12, name="Calibri")
ws1["A1"].alignment = Alignment(horizontal="center")

headers = ["Characteristic", f"Full Sample\n(N={N_total})", f"Pre-Period\n(N={N_pre})", f"Post-Period\n(N={N_post})"]
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 30

r = 3

def desc_section(ws, r, label, col, cats, df_full, df_pre, df_post):
    cell = ws.cell(row=r, column=1, value=label)
    cell.font = BOLD_FONT
    cell.fill = SUBHDR_FILL
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = SUBHDR_FILL
    r += 1
    for i, cat in enumerate(cats):
        n_f = (df_full[col] == cat).sum()
        n_pre = (df_pre[col] == cat).sum()
        n_post= (df_post[col] == cat).sum()
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        write_row(ws, r, [f"  {cat}",
                          pct(n_f, len(df_full)),
                          pct(n_pre, len(df_pre)),
                          pct(n_post, len(df_post))],
                  fill=fill, indent=1)
        r += 1
    return r

# Circuit
CIRCUITS = sorted(df["Circuit"].dropna().unique())
r = desc_section(ws1, r, "Circuit", "Circuit", CIRCUITS, df, pre, post)

# Period
r = desc_section(ws1, r, "Period", "Period (Pre/Post)", ["Pre", "Post"], df, pre, post)

# Adverse Action
ACTIONS = sorted(df["Primary Non-Economic Adverse Action"].dropna().unique())
r = desc_section(ws1, r, "Primary Adverse Action Type",
                 "Primary Non-Economic Adverse Action", ACTIONS, df, pre, post)

# Employer Type
EMP_TYPES = ["Government - Federal", "Government - Municipal", "Government - State", "Private"]
r = desc_section(ws1, r, "Employer Type", "Employer Type", EMP_TYPES, df, pre, post)

# Protected Characteristic
CHARS = sorted(df["Protected Characteristic(s)"].dropna().unique())
r = desc_section(ws1, r, "Protected Characteristic", "Protected Characteristic(s)", CHARS, df, pre, post)

# Appointing President Party
PARTIES = ["Democrat", "Republican"]
r = desc_section(ws1, r, "Appointing President Party", "Appointing President Party", PARTIES, df, pre, post)

# Motion Type
MOTIONS = sorted(df["Motion Type Decided"].dropna().unique())
r = desc_section(ws1, r, "Motion Type", "Motion Type Decided", MOTIONS, df, pre, post)

# Outcome vars
r = desc_section(ws1, r, "Court Finding on Harm", "Court Finding on Harm", HARM_ORDER, df, pre, post)
r = desc_section(ws1, r, "Primary Basis for Dismissal (standardised)",
                 "Dismissal_Clean", DISMISSAL_ORDER, df, pre, post)

add_note(ws1, r, "Note: Percentages may not sum to 100 due to rounding.")
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22

# ═════════════════════════════════════════════════════════════════════════════
# STEP 2 — CORE PRE/POST COMPARISON
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 2: Core Pre/Post Comparison ──")

ws2 = add_sheet("Table 2 — Pre Post Comparison")

ws2.merge_cells("A1:F1")
ws2["A1"] = "Table 2. Pre- vs. Post-Muldrow Comparison on Outcome Variables"
ws2["A1"].font = Font(bold=True, size=12, name="Calibri")
ws2["A1"].alignment = Alignment(horizontal="center")

hdrs2 = ["Outcome", f"Pre-Period\nN={N_pre}", "Pre %",
         f"Post-Period\nN={N_post}", "Post %", "χ²  p-value"]
for c, h in enumerate(hdrs2, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 30

r = 3

def two_way_section(ws, r, label, col, cats, df_pre, df_post, bold_rows=None):
    # Section header
    cell = ws.cell(row=r, column=1, value=label)
    cell.fill = SUBHDR_FILL
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for c in range(2, 7):
        ws.cell(row=r, column=c).fill = SUBHDR_FILL
    r += 1
    # Contingency table for chi2
    ct_pre  = df_pre[col].value_counts()
    ct_post = df_post[col].value_counts()
    ct = pd.DataFrame({"Pre": ct_pre, "Post": ct_post}).fillna(0)
    _, p_str = chi2_result(ct)
    first = True
    for i, cat in enumerate(cats):
        n_pre  = (df_pre[col]  == cat).sum()
        n_post = (df_post[col] == cat).sum()
        p_pre  = 100*n_pre/len(df_pre)  if len(df_pre)  else 0
        p_post = 100*n_post/len(df_post) if len(df_post) else 0
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        pval_cell = p_str if first else ""
        first = False
        vals = [f"  {cat}", n_pre, f"{p_pre:.1f}%", n_post, f"{p_post:.1f}%", pval_cell]
        write_row(ws, r, vals, fill=fill)
        r += 1
    return r

r = two_way_section(ws2, r, "Court Finding on Harm",
                    "Court Finding on Harm", HARM_ORDER, pre, post)
r = two_way_section(ws2, r, "Primary Basis for Dismissal",
                    "Dismissal_Clean", DISMISSAL_ORDER, pre, post)

add_note(ws2, r, "Note: χ² tests use Pearson chi-square without continuity correction.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F"], [35,12,10,12,10,16]):
    ws2.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 3 — THREE-WAY FILING DATE SUBGROUP
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 3: Three-Way Filing Date Subgroup ──")

ws3 = add_sheet("Table 3 — Filing Date Subgroup")

ws3.merge_cells("A1:H1")
ws3["A1"] = "Table 3. Outcome by Period and Filing Date Subgroup"
ws3["A1"].font = Font(bold=True, size=12, name="Calibri")
ws3["A1"].alignment = Alignment(horizontal="center")

post_prefiled  = df[df["Filing_Subgroup"] == "Post — Pre-Muldrow Filing"]
post_postfiled = df[df["Filing_Subgroup"] == "Post — Post-Muldrow Filing"]
N_ppre = len(post_prefiled)
N_ppost= len(post_postfiled)

hdrs3 = ["Outcome",
          f"Pre-Period\nN={N_pre}", "Pre %",
          f"Post — Pre-Muldrow Filing\nN={N_ppre}", "%",
          f"Post — Post-Muldrow Filing\nN={N_ppost}", "%",
          "χ²  p-value\n(3-way)"]
for c, h in enumerate(hdrs3, 1):
    cell = ws3.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[2].height = 35

r = 3

def three_way_section(ws, r, label, col, cats, d_pre, d_ppre, d_ppost):
    cell = ws.cell(row=r, column=1, value=label)
    cell.fill = SUBHDR_FILL
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = SUBHDR_FILL
    r += 1
    ct = pd.DataFrame({
        "Pre":        d_pre[col].value_counts(),
        "Post-Pre":   d_ppre[col].value_counts(),
        "Post-Post":  d_ppost[col].value_counts(),
    }).fillna(0)
    _, p_str = chi2_result(ct)
    first = True
    for i, cat in enumerate(cats):
        np_  = (d_pre[col]   == cat).sum()
        npp_ = (d_ppre[col]  == cat).sum()
        npo_ = (d_ppost[col] == cat).sum()
        pp_  = 100*np_/len(d_pre)   if len(d_pre)   else 0
        ppp_ = 100*npp_/len(d_ppre) if len(d_ppre)  else 0
        ppo_ = 100*npo_/len(d_ppost)if len(d_ppost) else 0
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        pval_cell = p_str if first else ""
        first = False
        vals = [f"  {cat}", np_, f"{pp_:.1f}%", npp_, f"{ppp_:.1f}%",
                npo_, f"{ppo_:.1f}%", pval_cell]
        write_row(ws, r, vals, fill=fill)
        r += 1
    return r

r = three_way_section(ws3, r, "Court Finding on Harm",
                      "Court Finding on Harm", HARM_ORDER,
                      pre, post_prefiled, post_postfiled)
r = three_way_section(ws3, r, "Primary Basis for Dismissal",
                      "Dismissal_Clean", DISMISSAL_ORDER,
                      pre, post_prefiled, post_postfiled)

add_note(ws3, r, "Note: χ² test is across all three groups.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F","G","H"], [35,12,9,18,9,20,9,16]):
    ws3.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 4 — CIRCUIT-LEVEL BREAKDOWN
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 4: Circuit-Level Breakdown ──")

ws4 = add_sheet("Table 4 — Circuit Breakdown")
ws4.merge_cells("A1:H1")
ws4["A1"] = "Table 4. Court Finding on Harm by Circuit — Pre vs. Post Muldrow"
ws4["A1"].font = Font(bold=True, size=12, name="Calibri")
ws4["A1"].alignment = Alignment(horizontal="center")

hdrs4 = ["Circuit",
          "Pre N", "Pre\nInsuff. %",
          "Post N", "Post\nInsuff. %", "Post\nSuff. %", "Post\nNot Reached %",
          "χ²  p-value"]
for c, h in enumerate(hdrs4, 1):
    cell = ws4.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws4.row_dimensions[2].height = 35

r = 3
circuits = ["First","Second","Fourth","Seventh","Eighth","Tenth","Eleventh"]
for i, circ in enumerate(circuits):
    d_pre_c  = pre[pre["Circuit"] == circ]
    d_post_c = post[post["Circuit"] == circ]
    def rate(d, cat): return 100*(d["Court Finding on Harm"]==cat).sum()/len(d) if len(d) else np.nan
    n_pre_c  = len(d_pre_c)
    n_post_c = len(d_post_c)
    insuff_pre  = rate(d_pre_c,  "Insufficient")
    insuff_post = rate(d_post_c, "Insufficient")
    suff_post   = rate(d_post_c, "Sufficient")
    nr_post     = rate(d_post_c, "Not Reached")
    ct = pd.DataFrame({
        "Pre":  d_pre_c["Court Finding on Harm"].value_counts(),
        "Post": d_post_c["Court Finding on Harm"].value_counts(),
    }).fillna(0)
    _, p_str = chi2_result(ct)
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    write_row(ws4, r, [circ, n_pre_c,
                       f"{insuff_pre:.1f}%" if not np.isnan(insuff_pre) else "—",
                       n_post_c,
                       f"{insuff_post:.1f}%" if not np.isnan(insuff_post) else "—",
                       f"{suff_post:.1f}%"   if not np.isnan(suff_post)  else "—",
                       f"{nr_post:.1f}%"     if not np.isnan(nr_post)    else "—",
                       p_str], fill=fill)
    r += 1

# Total row
def rate_tot(d, cat): return 100*(d["Court Finding on Harm"]==cat).sum()/len(d) if len(d) else np.nan
_, p_tot = chi2_result(pd.DataFrame({"Pre":pre["Court Finding on Harm"].value_counts(),
                                      "Post":post["Court Finding on Harm"].value_counts()}).fillna(0))
write_row(ws4, r, ["All Circuits", N_pre, f"{rate_tot(pre,'Insufficient'):.1f}%",
                    N_post,
                    f"{rate_tot(post,'Insufficient'):.1f}%",
                    f"{rate_tot(post,'Sufficient'):.1f}%",
                    f"{rate_tot(post,'Not Reached'):.1f}%",
                    p_tot], bold=True, fill=SUBHDR_FILL)
ws4.cell(row=r, column=1).font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
for c in range(2, 9):
    ws4.cell(row=r, column=c).font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
r += 1

add_note(ws4, r, "Note: χ² tests are pre vs. post within each circuit.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F","G","H"], [14,8,12,8,12,12,16,16]):
    ws4.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 5 — ADVERSE ACTION TYPE BREAKDOWN
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 5: Adverse Action Breakdown ──")

ws5 = add_sheet("Table 5 — Adverse Action")
ws5.merge_cells("A1:F1")
ws5["A1"] = "Table 5. Court Finding on Harm by Adverse Action Type"
ws5["A1"].font = Font(bold=True, size=12, name="Calibri")
ws5["A1"].alignment = Alignment(horizontal="center")

hdrs5 = ["Adverse Action Type",
          "Pre N", "Pre Insuff. %",
          "Post N", "Post Insuff. %",
          "χ²  p-value"]
for c, h in enumerate(hdrs5, 1):
    cell = ws5.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws5.row_dimensions[2].height = 30

r = 3
col_aa = "Primary Non-Economic Adverse Action"
actions_sorted = sorted(df[col_aa].dropna().unique())
for i, act in enumerate(actions_sorted):
    d_pre_a  = pre[pre[col_aa]  == act]
    d_post_a = post[post[col_aa] == act]
    n_pre_a  = len(d_pre_a)
    n_post_a = len(d_post_a)
    ins_pre  = 100*(d_pre_a["Court Finding on Harm"]=="Insufficient").sum()/n_pre_a  if n_pre_a  else np.nan
    ins_post = 100*(d_post_a["Court Finding on Harm"]=="Insufficient").sum()/n_post_a if n_post_a else np.nan
    ct = pd.DataFrame({"Pre": d_pre_a["Court Finding on Harm"].value_counts(),
                        "Post":d_post_a["Court Finding on Harm"].value_counts()}).fillna(0)
    _, p_str = chi2_result(ct)
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    write_row(ws5, r, [act, n_pre_a,
                        f"{ins_pre:.1f}%"  if not np.isnan(ins_pre)  else "—",
                        n_post_a,
                        f"{ins_post:.1f}%" if not np.isnan(ins_post) else "—",
                        p_str], fill=fill)
    r += 1

add_note(ws5, r, "Note: χ² tests are pre vs. post within each adverse action type.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F"], [30,8,14,8,14,16]):
    ws5.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 6 — EMPLOYER TYPE BREAKDOWN
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 6: Employer Type Breakdown ──")

ws6 = add_sheet("Table 6 — Employer Type")
ws6.merge_cells("A1:F1")
ws6["A1"] = "Table 6. Court Finding on Harm by Employer Type"
ws6["A1"].font = Font(bold=True, size=12, name="Calibri")
ws6["A1"].alignment = Alignment(horizontal="center")

for c, h in enumerate(["Employer Type","Pre N","Pre Insuff. %","Post N","Post Insuff. %","χ²  p-value"], 1):
    cell = ws6.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws6.row_dimensions[2].height = 30

r = 3
for i, emp in enumerate(EMP_TYPES):
    d_pre_e  = pre[pre["Employer Type"]  == emp]
    d_post_e = post[post["Employer Type"] == emp]
    n_pre_e  = len(d_pre_e)
    n_post_e = len(d_post_e)
    ins_pre  = 100*(d_pre_e["Court Finding on Harm"]=="Insufficient").sum()/n_pre_e   if n_pre_e  else np.nan
    ins_post = 100*(d_post_e["Court Finding on Harm"]=="Insufficient").sum()/n_post_e if n_post_e else np.nan
    ct = pd.DataFrame({"Pre": d_pre_e["Court Finding on Harm"].value_counts(),
                        "Post":d_post_e["Court Finding on Harm"].value_counts()}).fillna(0)
    _, p_str = chi2_result(ct)
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    write_row(ws6, r, [emp, n_pre_e,
                        f"{ins_pre:.1f}%"  if not np.isnan(ins_pre)  else "—",
                        n_post_e,
                        f"{ins_post:.1f}%" if not np.isnan(ins_post) else "—",
                        p_str], fill=fill)
    r += 1

add_note(ws6, r, "Note: χ² tests are pre vs. post within each employer type.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F"], [28,8,14,8,14,16]):
    ws6.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 7 — LOGISTIC REGRESSION
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 7: Logistic Regression ──")

ws7 = add_sheet("Table 7 — Logistic Regression")
ws7.merge_cells("A1:F1")
ws7["A1"] = "Table 7. Logistic Regression: Predictors of Harm Insufficient Finding"
ws7["A1"].font = Font(bold=True, size=12, name="Calibri")
ws7["A1"].alignment = Alignment(horizontal="center")

# Prepare regression dataset
rdf = df.copy()
rdf = rdf.dropna(subset=["Court Finding on Harm","Period (Pre/Post)","Circuit",
                           "Primary Non-Economic Adverse Action","Employer Type",
                           "Appointing President Party","Motion Type Decided"])
rdf["harm_insuff"] = (rdf["Court Finding on Harm"] == "Insufficient").astype(int)
rdf["post"]        = (rdf["Period (Pre/Post)"] == "Post").astype(int)

# Encode categoricals
for col in ["Circuit","Primary Non-Economic Adverse Action","Employer Type",
            "Appointing President Party","Motion Type Decided"]:
    rdf[col] = rdf[col].astype("category")

# Clean column names for formula
rdf.columns = [re.sub(r"[^A-Za-z0-9_]", "_", c) for c in rdf.columns]

formula1 = ("harm_insuff ~ post + C(Circuit) + "
            "C(Primary_Non_Economic_Adverse_Action) + "
            "C(Employer_Type) + "
            "C(Appointing_President_Party) + "
            "C(Motion_Type_Decided)")

model1 = smf.logit(formula1, data=rdf).fit(disp=False)

# Filing subgroup for spec 2 (post-period only for sub-group dummy)
rdf2 = rdf.copy()
rdf2["post_postfiled"] = ((rdf2["Period__Pre_Post_"] == "Post") &
                           (rdf2["Filing_Subgroup"] == "Post — Post-Muldrow Filing")).astype(int)

formula2 = ("harm_insuff ~ post + post_postfiled + C(Circuit) + "
            "C(Primary_Non_Economic_Adverse_Action) + "
            "C(Employer_Type) + "
            "C(Appointing_President_Party) + "
            "C(Motion_Type_Decided)")

model2 = smf.logit(formula2, data=rdf2).fit(disp=False)

def logit_rows(model, label):
    rows = []
    params = model.params
    conf   = model.conf_int()
    pvals  = model.pvalues
    for name in params.index:
        OR   = np.exp(params[name])
        lo   = np.exp(conf.loc[name, 0])
        hi   = np.exp(conf.loc[name, 1])
        p    = pvals[name]
        stars= sig_label(p)
        rows.append({
            "Variable": name,
            "OR":       f"{OR:.3f}",
            "95% CI":   f"[{lo:.3f}, {hi:.3f}]",
            "p-value":  f"{p:.3f}{stars}",
            "Spec":     label,
        })
    return rows

rows7  = logit_rows(model1, "Spec 1 (Base)")
rows7 += logit_rows(model2, "Spec 2 (+ Filing Subgroup)")

for c, h in enumerate(["Variable","Odds Ratio","95% CI","p-value","Specification"], 1):
    cell = ws7.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

r = 3
current_spec = None
for i, row in enumerate(rows7):
    if row["Spec"] != current_spec:
        current_spec = row["Spec"]
        cell = ws7.cell(row=r, column=1, value=current_spec)
        cell.fill = SUBHDR_FILL
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        for c in range(2, 6):
            ws7.cell(row=r, column=c).fill = SUBHDR_FILL
        r += 1
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    write_row(ws7, r, [row["Variable"], row["OR"], row["95% CI"], row["p-value"], row["Spec"]], fill=fill)
    r += 1

add_note(ws7, r,   f"Note: Spec 1 N={int(model1.nobs)}  |  Spec 2 N={int(model2.nobs)}.  "
                    "Reference categories: Circuit=Eighth, Adverse Action=Exclusion from Opportunities, "
                    "Employer=Government-Federal, Party=Democrat, Motion=12(b)(6).  "
                    "* p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E"], [55,14,22,14,28]):
    ws7.column_dimensions[col_ltr].width = w

print(f"  Model 1 pseudo-R²={model1.prsquared:.3f}  N={int(model1.nobs)}")
print(f"  Model 2 pseudo-R²={model2.prsquared:.3f}  N={int(model2.nobs)}")

# ═════════════════════════════════════════════════════════════════════════════
# STEP 8 — BINARY LOGISTIC REGRESSION: CAUSATION vs HARM INSUFFICIENT
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 8: Binary Logit (Causation vs Harm Insufficient) ──")

ws8 = add_sheet("Table 8 — Binary Logit")
ws8.merge_cells("A1:D1")
ws8["A1"] = "Table 8. Binary Logistic Regression: Causation Insufficient vs. Harm Insufficient"
ws8["A1"].font = Font(bold=True, size=12, name="Calibri")
ws8["A1"].alignment = Alignment(horizontal="center")

bdf = df[df["Dismissal_Clean"].isin(["Causation Insufficient", "Harm Insufficient"])].copy()
bdf["causation"] = (bdf["Dismissal_Clean"] == "Causation Insufficient").astype(int)
bdf["post"]      = (bdf["Period (Pre/Post)"] == "Post").astype(int)
bdf = bdf.dropna(subset=["Circuit", "Employer Type",
                           "Appointing President Party", "Motion Type Decided"])
bdf.columns = [re.sub(r"[^A-Za-z0-9_]", "_", c) for c in bdf.columns]

bformula = ("causation ~ post + C(Circuit) + "
            "C(Employer_Type) + "
            "C(Appointing_President_Party) + "
            "C(Motion_Type_Decided)")

try:
    bmodel = smf.logit(bformula, data=bdf).fit(method="bfgs", maxiter=1000, disp=False)

    for c, h in enumerate(["Variable", "Odds Ratio", "95% CI", "p-value"], 1):
        cell = ws8.cell(row=2, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 3
    params = bmodel.params
    conf   = bmodel.conf_int()
    pvals  = bmodel.pvalues
    for i, name in enumerate(params.index):
        OR    = np.exp(params[name])
        lo    = np.exp(conf.loc[name, 0])
        hi    = np.exp(conf.loc[name, 1])
        p     = pvals[name]
        stars = sig_label(p)
        fill  = ALT_FILL if i % 2 == 0 else WHITE_FILL
        write_row(ws8, r, [name, f"{OR:.3f}", f"[{lo:.3f}, {hi:.3f}]", f"{p:.3f}{stars}"], fill=fill)
        r += 1

    add_note(ws8, r,
             f"Note: N={int(bmodel.nobs)}. Outcome=1 if Causation Insufficient, 0 if Harm Insufficient. "
             "Sample restricted to dismissed cases coded as either Causation Insufficient or Harm Insufficient. "
             "Reference categories: Circuit=Eighth, Employer=Government-Federal, Party=Democrat, Motion=12(b)(6). "
             f"Pseudo-R²={bmodel.prsquared:.3f}.  * p<0.05  ** p<0.01  *** p<0.001")

    if "post" in params.index:
        print(f"  OR_post={np.exp(params['post']):.3f}  p={pvals['post']:.4f}")

except Exception as e:
    ws8.cell(row=3, column=1, value=f"Model failed: {e}")
    print(f"  Binary logit error: {e}")

for col_ltr, w in zip(["A","B","C","D"], [55,14,22,14]):
    ws8.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# STEP 9 — ROBUSTNESS CHECKS
# ═════════════════════════════════════════════════════════════════════════════
print("── Step 9: Robustness Checks ──")

ws9 = add_sheet("Table 9 — Robustness Checks")
ws9.merge_cells("A1:H1")
ws9["A1"] = "Table 9. Robustness Checks — Core Pre/Post Comparison on Restricted Samples"
ws9["A1"].font = Font(bold=True, size=12, name="Calibri")
ws9["A1"].alignment = Alignment(horizontal="center")

hdrs9 = ["Sample Restriction",
          "Pre N", "Pre Insuff. %",
          "Post N", "Post Insuff. %",
          "Pre Survived %", "Post Survived %",
          "χ²  p-value"]
for c, h in enumerate(hdrs9, 1):
    cell = ws9.cell(row=2, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws9.row_dimensions[2].height = 35

r = 3
samples = [
    ("Full sample (baseline)",
     df[df["Period (Pre/Post)"]=="Pre"],
     df[df["Period (Pre/Post)"]=="Post"]),
    ("Excluding Uncertain confidence",
     df[(df["Period (Pre/Post)"]=="Pre") & (df["Dismissal Basis Confidence"]!="Uncertain")],
     df[(df["Period (Pre/Post)"]=="Post") & (df["Dismissal Basis Confidence"]!="Uncertain")]),
    ("Excluding Multiple Bases",
     df[(df["Period (Pre/Post)"]=="Pre") & (df["Primary Basis for Dismissal"]!="Multiple Bases")],
     df[(df["Period (Pre/Post)"]=="Post") & (df["Primary Basis for Dismissal"]!="Multiple Bases")]),
    ("Excluding Fourth Circuit",
     df[(df["Period (Pre/Post)"]=="Pre") & (df["Circuit"]!="Fourth")],
     df[(df["Period (Pre/Post)"]=="Post") & (df["Circuit"]!="Fourth")]),
    ("Post — post-Muldrow filing only",
     df[df["Period (Pre/Post)"]=="Pre"],
     df[df["Filing_Subgroup"]=="Post — Post-Muldrow Filing"]),
]

for i, (label, d_pre_r, d_post_r) in enumerate(samples):
    n_pre_r  = len(d_pre_r)
    n_post_r = len(d_post_r)
    def rate_r(d, cat):
        return 100*(d["Court Finding on Harm"]==cat).sum()/len(d) if len(d) else np.nan
    ins_pre_r  = rate_r(d_pre_r,  "Insufficient")
    ins_post_r = rate_r(d_post_r, "Insufficient")
    surv_pre_r = 100*(d_pre_r["Dismissal_Clean"]=="N/A — Plaintiff Survived").sum()/n_pre_r  if n_pre_r  else np.nan
    surv_post_r= 100*(d_post_r["Dismissal_Clean"]=="N/A — Plaintiff Survived").sum()/n_post_r if n_post_r else np.nan
    ct = pd.DataFrame({"Pre": d_pre_r["Court Finding on Harm"].value_counts(),
                        "Post":d_post_r["Court Finding on Harm"].value_counts()}).fillna(0)
    _, p_str = chi2_result(ct)
    fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
    write_row(ws9, r, [label, n_pre_r,
                        f"{ins_pre_r:.1f}%"  if not np.isnan(ins_pre_r)  else "—",
                        n_post_r,
                        f"{ins_post_r:.1f}%" if not np.isnan(ins_post_r) else "—",
                        f"{surv_pre_r:.1f}%" if not np.isnan(surv_pre_r) else "—",
                        f"{surv_post_r:.1f}%"if not np.isnan(surv_post_r)else "—",
                        p_str], fill=fill)
    r += 1

add_note(ws9, r, "Note: χ² tests compare Court Finding on Harm across all categories, Pre vs. Post.  * p<0.05  ** p<0.01  *** p<0.001")
for col_ltr, w in zip(["A","B","C","D","E","F","G","H"], [40,8,14,8,14,14,14,16]):
    ws9.column_dimensions[col_ltr].width = w

# ═════════════════════════════════════════════════════════════════════════════
# SAVE EXCEL
# ═════════════════════════════════════════════════════════════════════════════
wb.save(XLOUT)
print(f"\nExcel workbook saved: {XLOUT}")

# ═════════════════════════════════════════════════════════════════════════════
# STEP 10 — VISUALIZATIONS
# ═════════════════════════════════════════════════════════════════════════════
print("\n── Step 10: Visualizations ──")

def save_fig(fig, name):
    path = os.path.join(OUTDIR, f"{name}.png")
    fig.savefig(path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {path}")
    return path

# ── Fig 1: Pre vs Post — Court Finding on Harm ───────────────────────────────
fig, ax = plt.subplots(figsize=(8, 5))
counts = pd.DataFrame({
    "Pre":  pre["Court Finding on Harm"].value_counts().reindex(HARM_ORDER, fill_value=0),
    "Post": post["Court Finding on Harm"].value_counts().reindex(HARM_ORDER, fill_value=0),
})
pcts = counts.div(counts.sum()) * 100

x = np.arange(len(HARM_ORDER))
w = 0.35
bars_pre  = ax.bar(x - w/2, pcts["Pre"],  w, label=f"Pre-Period (N={N_pre})",  color=BLUE,   edgecolor="white", linewidth=0.5)
bars_post = ax.bar(x + w/2, pcts["Post"], w, label=f"Post-Period (N={N_post})", color=ORANGE, edgecolor="white", linewidth=0.5)

for bars in [bars_pre, bars_post]:
    for bar in bars:
        h = bar.get_height()
        if h > 2:
            ax.text(bar.get_x() + bar.get_width()/2, h + 0.5, f"{h:.1f}%",
                    ha="center", va="bottom", fontsize=8.5, fontweight="bold")

ax.set_xlabel("Court Finding on Harm", labelpad=6)
ax.set_ylabel("Percentage of Cases (%)", labelpad=6)
ax.set_title("Figure 1. Court Finding on Harm: Pre- vs. Post-Muldrow", pad=10)
ax.set_xticks(x)
ax.set_xticklabels(HARM_ORDER)
ax.set_ylim(0, max(pcts.values.max() * 1.18, 10))
ax.legend(frameon=False)
ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
ax.set_axisbelow(True)
save_fig(fig, "fig1_harm_finding_pre_post")

# ── Fig 2: Pre vs Post — Primary Basis for Dismissal (3 categories) ──────────
FIG2_CATS = ["Harm Insufficient", "Causation Insufficient", "N/A — Plaintiff Survived"]
pre_fig2  = pre[pre["Dismissal_Clean"].isin(FIG2_CATS)]
post_fig2 = post[post["Dismissal_Clean"].isin(FIG2_CATS)]

fig, ax = plt.subplots(figsize=(8, 5))
counts2 = pd.DataFrame({
    "Pre":  pre_fig2["Dismissal_Clean"].value_counts().reindex(FIG2_CATS, fill_value=0),
    "Post": post_fig2["Dismissal_Clean"].value_counts().reindex(FIG2_CATS, fill_value=0),
})
pcts2 = counts2.div(counts2.sum()) * 100

x2 = np.arange(len(FIG2_CATS))
bars_pre2  = ax.bar(x2 - w/2, pcts2["Pre"],  w, label=f"Pre-Period (N={len(pre_fig2)})",  color=BLUE,   edgecolor="white")
bars_post2 = ax.bar(x2 + w/2, pcts2["Post"], w, label=f"Post-Period (N={len(post_fig2)})", color=ORANGE, edgecolor="white")

for bars in [bars_pre2, bars_post2]:
    for bar in bars:
        h = bar.get_height()
        if h > 2:
            ax.text(bar.get_x() + bar.get_width()/2, h + 0.5, f"{h:.1f}%",
                    ha="center", va="bottom", fontsize=8.5, fontweight="bold")

short_labels = ["Harm\nInsuff.", "Causation\nInsuff.", "N/A —\nSurvived"]
ax.set_xticks(x2)
ax.set_xticklabels(short_labels, fontsize=10)
ax.set_xlabel("Primary Basis for Dismissal", labelpad=6)
ax.set_ylabel("Percentage of Cases with Recorded Basis (%)", labelpad=6)
ax.set_title("Figure 2. Primary Basis for Dismissal: Pre- vs. Post-Muldrow", pad=10)
ax.set_ylim(0, max(pcts2.values.max() * 1.18, 10))
ax.legend(frameon=False)
ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
ax.set_axisbelow(True)
save_fig(fig, "fig2_dismissal_basis_pre_post")

# ── Fig 3: Three-Way Filing Date Comparison ───────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 5))
groups = ["Pre-Period", "Post — Pre-Muldrow\nFiling", "Post — Post-Muldrow\nFiling"]
dsets  = [pre, post_prefiled, post_postfiled]
counts3 = pd.DataFrame({
    g: d["Court Finding on Harm"].value_counts().reindex(HARM_ORDER, fill_value=0)
    for g, d in zip(groups, dsets)
})
pcts3 = counts3.div(counts3.sum()) * 100

x3  = np.arange(len(groups))
bw  = 0.20
colors3 = PAL_4
for k, cat in enumerate(HARM_ORDER):
    offset = (k - 1.5) * bw
    vals = pcts3.loc[cat].values
    bars = ax.bar(x3 + offset, vals, bw, label=cat,
                  color=colors3[k], edgecolor="white", linewidth=0.5)
    for bar in bars:
        h = bar.get_height()
        if h > 3:
            ax.text(bar.get_x() + bar.get_width()/2, h + 0.4, f"{h:.0f}%",
                    ha="center", va="bottom", fontsize=7, fontweight="bold")

ax.set_xticks(x3)
ax.set_xticklabels(groups, fontsize=9)
ax.set_xlabel("Group", labelpad=6)
ax.set_ylabel("Percentage of Cases (%)", labelpad=6)
ax.set_title("Figure 3. Court Finding on Harm by Filing Date Subgroup", pad=10)
ax.set_ylim(0, max(pcts3.values.max() * 1.20, 10))
ax.legend(frameon=False, fontsize=8.5, ncol=2)
ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
ax.set_axisbelow(True)
# annotate N
for xi, d in zip(x3, dsets):
    ax.text(xi, -5, f"N={len(d)}", ha="center", fontsize=8, color="grey", transform=ax.transData)
save_fig(fig, "fig3_three_way_filing")

# ── Fig 4: Post-Period Harm Insufficient Rate by Circuit (grouped bar) ────────
fig, ax = plt.subplots(figsize=(9, 5))
circ_pre_insuff  = []
circ_post_insuff = []
circ_n_pre       = []
circ_n_post      = []
for circ in circuits:
    dp = pre[pre["Circuit"] == circ]
    dq = post[post["Circuit"] == circ]
    circ_n_pre.append(len(dp))
    circ_n_post.append(len(dq))
    circ_pre_insuff.append(100*(dp["Court Finding on Harm"]=="Insufficient").sum()/len(dp) if len(dp) else 0)
    circ_post_insuff.append(100*(dq["Court Finding on Harm"]=="Insufficient").sum()/len(dq) if len(dq) else 0)

xc = np.arange(len(circuits))
bw = 0.38
ax.bar(xc - bw/2, circ_pre_insuff,  bw, label="Pre-Period",  color=BLUE,   edgecolor="white")
ax.bar(xc + bw/2, circ_post_insuff, bw, label="Post-Period", color=ORANGE, edgecolor="white")
for xi, (pre_v, post_v, np_, nq_) in enumerate(zip(circ_pre_insuff, circ_post_insuff, circ_n_pre, circ_n_post)):
    ax.text(xi - bw/2, pre_v + 1, f"{pre_v:.0f}%\n(N={np_})",  ha="center", fontsize=7.5)
    ax.text(xi + bw/2, post_v+ 1, f"{post_v:.0f}%\n(N={nq_})", ha="center", fontsize=7.5)

circ_labels = [f"First*" if c == "First" else c for c in circuits]
ax.set_xticks(xc)
ax.set_xticklabels(circ_labels, fontsize=9)
ax.set_xlabel("Circuit", labelpad=6)
ax.set_ylabel("Harm Insufficient Rate (%)", labelpad=6)
ax.set_title("Figure 4. Harm Insufficient Rate by Circuit: Pre vs. Post Muldrow", pad=10)
ax.set_ylim(0, 105)
ax.legend(frameon=False)
ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
ax.set_axisbelow(True)
ax.text(0.01, -0.12, "* First Circuit: n=1 pre-period", transform=ax.transAxes,
        fontsize=8, color="grey", style="italic")
save_fig(fig, "fig4_circuit_harm_insufficient")

# ── Fig 5: Harm Insufficient Rate by Adverse Action Type (≥10 pre-period) ─────
col_aa = "Primary Non-Economic Adverse Action"
# Restrict to categories with ≥10 pre-period cases
fig5_actions = [act for act in actions_sorted
                if (pre[col_aa] == act).sum() >= 10]

aa_pre_insuff  = []
aa_post_insuff = []
aa_n_pre, aa_n_post = [], []
for act in fig5_actions:
    dp = pre[pre[col_aa]  == act]
    dq = post[post[col_aa] == act]
    aa_n_pre.append(len(dp))
    aa_n_post.append(len(dq))
    aa_pre_insuff.append(100*(dp["Court Finding on Harm"]=="Insufficient").sum()/len(dp) if len(dp) else 0)
    aa_post_insuff.append(100*(dq["Court Finding on Harm"]=="Insufficient").sum()/len(dq) if len(dq) else 0)

fig, ax = plt.subplots(figsize=(11, 5))
xa = np.arange(len(fig5_actions))
bwa = 0.38
ax.bar(xa - bwa/2, aa_pre_insuff,  bwa, label="Pre-Period",  color=BLUE,   edgecolor="white")
ax.bar(xa + bwa/2, aa_post_insuff, bwa, label="Post-Period", color=ORANGE, edgecolor="white")

short_aa = [a.replace(" / ", "/\n") for a in fig5_actions]
ax.set_xticks(xa)
ax.set_xticklabels(short_aa, fontsize=11, rotation=45, ha="right")
ax.set_xlabel("Adverse Action Type", labelpad=6)
ax.set_ylabel("Harm Insufficient Rate (%)", labelpad=6)
ax.set_title("Figure 5. Harm Insufficient Rate by Adverse Action Type: Pre vs. Post Muldrow", pad=10)
ax.set_ylim(0, 115)
for xi, (pre_v, post_v, np_, nq_) in enumerate(zip(aa_pre_insuff, aa_post_insuff, aa_n_pre, aa_n_post)):
    if np_ > 0:
        ax.text(xi - bwa/2, pre_v + 1.5, f"N={np_}", ha="center", fontsize=7, color=BLUE)
    if nq_ > 0:
        ax.text(xi + bwa/2, post_v + 1.5, f"N={nq_}", ha="center", fontsize=7, color=ORANGE)
ax.legend(frameon=False)
ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
ax.set_axisbelow(True)
plt.tight_layout()
save_fig(fig, "fig5_adverse_action_harm_insufficient")

print("\n═══ All steps complete ═══")
print(f"  Tables: {XLOUT}")
print(f"  Figures: {OUTDIR}/fig1–5.png")
